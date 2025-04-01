import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def save_filtered_data(df, output_csv, excel_path, sheet_name="FilteredData"):
    df.to_csv(output_csv, index=False)

    book = openpyxl.load_workbook(excel_path)
    if sheet_name in book.sheetnames:
        book.remove(book[sheet_name])
    sheet = book.create_sheet(title=sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    book.save(excel_path)